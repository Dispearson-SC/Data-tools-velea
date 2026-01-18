import pandas as pd
import re
from datetime import datetime, timedelta
from .utils import leer_archivo_base
import openpyxl

PALABRAS_CLAVE = {
    'TRA_END': ['TOTALES TRA', 'TOTALES TRADICIONAL'],
    'HP_START': ['TOTALES HP', 'TOTALES HOJA'],
    'BORR_START': ['TOTAL BORR', 'TOTALES BORR', 'TOTALES BORRACHO'],
    'BORR_END': ['TOTAL ELOTE', 'ELOTE']
}

def calcular_fecha_inicio(numero_semana):
    try:
        num = int(numero_semana)
        fecha_base = datetime(2025, 12, 29) # Lunes de Sem #1
        fecha_inicio = fecha_base + timedelta(days=(num - 1) * 7)
        return fecha_inicio
    except ValueError:
        return None

def limpiar_texto(texto):
    if pd.isna(texto): return ""
    return re.sub(r'\s+', ' ', str(texto)).strip().upper()

def buscar_fila_clave(df, lista_palabras, n_cols=5):
    for col_idx in range(n_cols):
        if col_idx >= len(df.columns): break
        col_str = df.iloc[:, col_idx].apply(limpiar_texto)
        for palabra in lista_palabras:
            matches = col_str[col_str.str.contains(palabra, na=False)]
            if not matches.empty:
                return matches.index[0]
    return None

def identificar_guiso(texto):
    t = limpiar_texto(texto).replace('.', '')
    mapeo = {
        'F': 'FRIJOL', 'FRIJOL': 'FRIJOL', 'FRIJO': 'FRIJOL',
        'P': 'POLLO', 'POLLO': 'POLLO', 'POLL': 'POLLO',
        'Q': 'QUESO', 'QUESO': 'QUESO', 'QUES': 'QUESO',
        'PCO': 'PCO', 'PUERCO': 'PCO', 'CERDO': 'PCO',
        'SV': 'SALSA VERDE', 'VERDE': 'SALSA VERDE', 'SALSA': 'SALSA VERDE',
        'DULCE': 'DULCE', 'RAJAS': 'RAJAS', 'CHICHARRON': 'CHICHARRON',
        'PICADILLO': 'PICADILLO'
    }
    return mapeo.get(t, t)

def interpretar_nota(texto_nota, total_celda):
    if not texto_nota: return None
    texto = texto_nota.lower().replace('\n', ' ')
    resultados = {}
    patrones = [r'(\d+)\s*(?:de)?\s*([a-z\.]+)', r'([a-z\.]+)\s*:?\s*(\d+)']

    for p in patrones:
        for match in re.findall(p, texto):
            try:
                cant = float(match[0])
                palabra = match[1]
            except:
                try:
                    cant = float(match[1])
                    palabra = match[0]
                except: continue

            guiso = identificar_guiso(palabra)
            if guiso and guiso not in ['DE', 'Y', 'EL', 'LA', 'CAJAS', 'TOTAL']:
                resultados[guiso] = resultados.get(guiso, 0) + cant

    if resultados and abs(sum(resultados.values()) - total_celda) <= 2:
        return resultados
    return None

def procesar_produccion(files_content: list):
    datos_totales = []

    for filename, content in files_content:
        # Needs to be Excel for this logic
        if not filename.endswith(('.xlsx', '.xls')):
            continue
        
        try:
            # We need openpyxl for comments, so we use load_workbook logic if possible
            # But leer_archivo_base returns ExcelFile or DF. 
            # We need to adapt logic to work with bytes -> openpyxl
            import io
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            # We need comments, data_only=True removes formulas but keeps comments? 
            # Actually data_only=True might lose comments in some versions, but usually comments are separate.
            # But to get comments we need the cell object.
            # Let's check the original script: it used `openpyxl.load_workbook(archivo, data_only=True)`
            # And accessed `celda.comment`. So it works.
            
            for hoja in wb.sheetnames:
                match = re.search(r'SEM.*?(\d+)', hoja.upper())
                if not match:
                    if hoja.strip().isdigit():
                        num_sem = int(hoja)
                    else:
                        continue
                else:
                    num_sem = int(match.group(1))

                fecha_inicio = calcular_fecha_inicio(num_sem)
                ws = wb[hoja]
                data = list(ws.values)
                if not data: continue
                df = pd.DataFrame(data)

                idx_tra_end = buscar_fila_clave(df, PALABRAS_CLAVE['TRA_END'])
                idx_hp_start = buscar_fila_clave(df, PALABRAS_CLAVE['HP_START'])
                idx_borr_start = buscar_fila_clave(df, PALABRAS_CLAVE['BORR_START'])
                idx_borr_end = buscar_fila_clave(df, PALABRAS_CLAVE['BORR_END']) or len(df)

                if None in [idx_tra_end, idx_hp_start, idx_borr_start]:
                    continue

                bloques = [
                    (range(3, idx_tra_end), "Tradicional"),
                    (range(idx_hp_start+1, idx_borr_start), "Hoja de Platano (HP)"),
                    (range(idx_borr_start+1, idx_borr_end), "Borracho")
                ]

                for rango, tipo in bloques:
                    for r_idx in rango:
                        fila = df.iloc[r_idx]
                        nombre = fila[0]
                        if pd.isna(nombre) or "TOTAL" in str(nombre).upper(): continue

                        for i in range(7): # 7 días
                            c_guiso = 2 + (i * 3)
                            c_cant = 3 + (i * 3)
                            if c_cant >= len(fila): continue

                            try: cant = float(fila[c_cant])
                            except: cant = 0

                            if cant > 0:
                                guiso_raw = limpiar_texto(fila[c_guiso])
                                guiso_final = guiso_raw if guiso_raw else "DESCONOCIDO"
                                registrado = False

                                # Manejo de mezclas con notas
                                if '/' in guiso_raw:
                                    # Access cell for comment
                                    # openpyxl is 1-indexed for rows/cols
                                    celda = ws.cell(row=r_idx+1, column=c_guiso+1)
                                    if celda.comment:
                                        desglose = interpretar_nota(celda.comment.text, cant)
                                        if desglose:
                                            registrado = True
                                            for g, c in desglose.items():
                                                datos_totales.append({
                                                    'Fecha': fecha_inicio + timedelta(days=i),
                                                    'Tipo_Tamal': tipo, 'Guiso': g, 'Cantidad': c
                                                })

                                    if not registrado:
                                        partes = [p for p in guiso_raw.split('/') if p.strip()]
                                        if partes:
                                            split = round(cant / len(partes), 2)
                                            for p in partes:
                                                datos_totales.append({
                                                    'Fecha': fecha_inicio + timedelta(days=i),
                                                    'Tipo_Tamal': tipo, 'Guiso': identificar_guiso(p), 'Cantidad': split
                                                })
                                            registrado = True

                                if not registrado:
                                    datos_totales.append({
                                        'Fecha': fecha_inicio + timedelta(days=i),
                                        'Tipo_Tamal': tipo,
                                        'Guiso': identificar_guiso(guiso_final),
                                        'Cantidad': cant
                                    })

        except Exception as e:
            print(f"Error processing production file {filename}: {e}")
            continue

    if datos_totales:
        df_fin = pd.DataFrame(datos_totales)
        df_agrupado = df_fin.groupby(['Fecha', 'Tipo_Tamal', 'Guiso'], as_index=False)['Cantidad'].sum()
        df_agrupado.sort_values(by=['Fecha', 'Tipo_Tamal', 'Guiso'], inplace=True)
        return df_agrupado
    
    return pd.DataFrame()
